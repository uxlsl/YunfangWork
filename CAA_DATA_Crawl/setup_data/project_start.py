#coding=utf-8
import os,re,time,sys,json,copy,urllib
from redis import Redis
from  openpyxl.reader.excel import load_workbook 
from sqlalchemy import *
from sqlalchemy.orm import *
reload(sys)
sys.setdefaultencoding( "utf-8" )



city_name_col_index = 2
database_name_col_index = 3
server_address_col_inex = 4
user_name = 'sa'
passwd = 'gh001'
port = '1433'
data_source_path = '/home/chiufung/xxjr_info/source_info/info_data.xlsx'
push_key = 'xxjr:seed_info'
r = Redis(host='192.168.6.10')

wb = load_workbook(filename=data_source_path)

for sheet_name in wb.get_sheet_names():
    print sheet_name
    ws = wb.get_sheet_by_name(sheet_name)
    nrows = ws.max_row
    for row_index in xrange(2,nrows+1):
        try:
            city_info = ws.cell(row=row_index,column=city_name_col_index).value.split('/')
            database_name = ws.cell(row=row_index,column=database_name_col_index).value
            server_address = ws.cell(row=row_index,column=server_address_col_inex).value
            if city_info and database_name and server_address:
                if len(city_info)>1:
                    city_name = city_info[1]
                    mssql_path = 'mssql+pymssql://{user}:{passwd}@{address}:{port}/{dbname}'.\
                                    format(user=user_name,passwd=passwd,address=server_address,port=port,dbname=database_name)
                    engine = create_engine(mssql_path)
                    metadata = MetaData(engine)
                    RANameInfo_tab = Table('RANameInfo', metadata, autoload=True)
                    Session = sessionmaker(bind=engine)
                    s = Session()
                    for instance in s.query(RANameInfo_tab):
                        try:
                            key_word = instance.NameInSource
                            city = city_name
                            input_data = '{"cityName":"%s","keyWord":"%s"}'%(city,key_word)
                            r.sadd(push_key,input_data)
                            print input_data
                        except Exception as e:
                             print Exception,":",e
                else:
                    continue
            else:
                continue
        except Exception as e:
            print Exception,":",e

            






    
    